from __future__ import annotations
import io, json, hashlib
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models import CalculationRun

def _run_hash(run: CalculationRun) -> str:
    payload = {
        "id": run.id,
        "run_type": run.run_type,
        "total_tco2e": run.total_tco2e,
        "details": run.details,
        "created_at": run.created_at.isoformat(),
    }
    b = json.dumps(payload, sort_keys=True).encode("utf-8")
    return hashlib.sha256(b).hexdigest()

def export_run_pdf(db: Session, run_id: int) -> bytes:
    run = db.query(CalculationRun).filter(CalculationRun.id == run_id).one_or_none()
    if not run:
        raise ValueError("Run not found")
    sha = _run_hash(run)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    y = h - 50

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Carbon Platform Report (Run)")
    y -= 20
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Run ID: {run.id}   Type: {run.run_type}   Created: {run.created_at.isoformat()}")
    y -= 14
    c.drawString(50, y, f"Total: {run.total_tco2e:.6f} tCO2e ({run.total_kgco2e:.2f} kgCO2e)")
    y -= 14
    c.drawString(50, y, f"SHA256: {sha}")
    y -= 20

    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y, "Rows (first 40):")
    y -= 14
    c.setFont("Helvetica", 9)

    rows = (run.details or {}).get("rows") or []
    for row in rows[:40]:
        line = f"- activity_id={row.get('activity_id')} ef={row.get('ef_key')} kgCO2e={float(row.get('kgco2e',0)):.4f}"
        c.drawString(60, y, line[:120])
        y -= 12
        if y < 80:
            c.showPage()
            y = h - 50
            c.setFont("Helvetica", 9)

    c.showPage()
    c.save()
    return buf.getvalue()

def export_run_excel(db: Session, run_id: int) -> bytes:
    run = db.query(CalculationRun).filter(CalculationRun.id == run_id).one_or_none()
    if not run:
        raise ValueError("Run not found")
    sha = _run_hash(run)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Run ID", run.id])
    ws.append(["Run Type", run.run_type])
    ws.append(["Created", run.created_at.isoformat()])
    ws.append(["Total tCO2e", run.total_tco2e])
    ws.append(["Total kgCO2e", run.total_kgco2e])
    ws.append(["SHA256", sha])

    ws2 = wb.create_sheet("Rows")
    ws2.append(["activity_id","activity_name","ef_key","kgco2e","inputs_json","trace_json"])
    for row in ((run.details or {}).get("rows") or []):
        ws2.append([
            row.get("activity_id"),
            row.get("activity_name"),
            row.get("ef_key"),
            row.get("kgco2e"),
            json.dumps(row.get("inputs") or {}, ensure_ascii=False),
            json.dumps(row.get("trace") or {}, ensure_ascii=False),
        ])
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
